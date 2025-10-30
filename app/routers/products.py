from typing import List
from fastapi import APIRouter, Depends, status, UploadFile, File
from fastapi.responses import StreamingResponse
from pydantic import ValidationError
from app.core.exceptions import BadRequestException
from logger import logger

from io import BytesIO
from openpyxl import Workbook, load_workbook

from app.schemas.product import (
    ProductCreate,
    ProductImageCreate,
    ProductImportRow,
    ProductUpdate,
    ProductRead,
    ProductList,
)
from app.services.product import ProductService
from app.utils.deps import get_product_service

router = APIRouter(prefix="/products", tags=["Products"])


@router.post("/", response_model=ProductRead, status_code=status.HTTP_201_CREATED)
async def create_product(
    product: ProductCreate, service: ProductService = Depends(get_product_service)
):
    db_product = await service.create_product(product)
    logger.info(f"Product created: {db_product.name}")
    return db_product


@router.get("/", response_model=ProductList)
async def get_products(
    skip: int = 0,
    limit: int = 10,
    category_id: int = None,
    category: str = None,
    service: ProductService = Depends(get_product_service),
):

    filters = {
        k: v
        for k, v in {
            "category_id": category_id,
            "category": category,
        }.items()
        if v is not None
    }

    return await service.get_products(skip, limit, **filters)


@router.post("/import")
async def import_products(
    file: UploadFile = File(...), service: ProductService = Depends(get_product_service)
):
    if not file.filename.endswith(".xlsx"):
        raise BadRequestException("Підтримується лише формат .xlsx")

    workbook = load_workbook(filename=file.file)
    sheet = workbook.active

    # Очікувані колонки (у тому ж порядку, що в Excel)
    expected_columns = [
        "sku",
        "name",
        "description",
        "category_id",
        "base_price",
        "stock_quantity",
        "image_url",
    ]
    header = [cell.value for cell in sheet[1]]

    if header != expected_columns:
        raise BadRequestException(
            f"Неправильні заголовки стовпців. Очікується: {', '.join(expected_columns)}",
        )

    # Збір даних з Excel
    rows_data: List[ProductImportRow] = []
    errors = []
    seen_skus = set()

    for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        try:
            row_dict = dict(zip(expected_columns, row))

            # Перевірка на дублікати SKU у файлі
            if row_dict["sku"] in seen_skus:
                raise ValueError(f"SKU '{row_dict['sku']}' повторюється у файлі")
            seen_skus.add(row_dict["sku"])

            validated = ProductImportRow(**row_dict)
            rows_data.append(validated)

        except ValidationError as e:
            errors.append({"row": idx, "error": e.errors()})
        except ValueError as e:
            errors.append({"row": idx, "error": str(e)})

    # Якщо є помилки — не додаємо нічого в базу
    if errors:
        raise BadRequestException(
            detail={"message": "Помилка валідації даних", "errors": errors},
        )

    created_count = 0
    updated_count = 0

    for row in rows_data:
        existing = await service.get_product_by_sku(row.sku)

        if existing:
            await service.update_product(
                existing.id,
                ProductUpdate(
                    name=row.name,
                    description=row.description,
                    sku=row.sku,
                    category_id=row.category_id,
                    base_price=row.base_price,
                    stock_quantity=row.stock_quantity,
                ),
            )
            updated_count += 1
        else:
            await service.create_product(
                ProductCreate(
                    name=row.name,
                    description=row.description,
                    sku=row.sku,
                    category_id=row.category_id,
                    base_price=row.base_price,
                    stock_quantity=row.stock_quantity,
                    images=(
                        [ProductImageCreate(image_url=row.image_url, is_main=True)]
                        if row.image_url
                        else []
                    ),
                )
            )
            created_count += 1

    return {
        "message": f"Імпортовано {created_count} продуктів, оновлено {updated_count}",
        "total": created_count + updated_count,
    }


@router.get("/export")
async def export_products(service: ProductService = Depends(get_product_service)):
    products = await service.export_products()

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(
        [
            "sku",
            "name",
            "description",
            "category_id",
            "base_price",
            "stock_quantity",
            "image_url",
        ]
    )
    for p in products:
        sheet.append(
            (
                p.sku,
                p.name,
                p.description,
                p.category_id,
                p.base_price,
                p.stock_quantity,
                p.images[0].image_url if p.images else "",
            )
        )

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)

    headers = {"Content-Disposition": 'attachment; filename="products.xlsx"'}
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@router.get("/{product_id}", response_model=ProductRead)
async def get_product(
    product_id: int, service: ProductService = Depends(get_product_service)
):
    product = await service.get_product(product_id)
    return product


@router.get("/slug/{slug}", response_model=ProductRead)
async def get_product_by_slug(
    slug: str, service: ProductService = Depends(get_product_service)
):
    product = await service.get_product_by_slug(slug)
    return product


@router.put("/{product_id}", response_model=ProductRead)
async def update_product(
    product_id: int,
    product_data: ProductUpdate,
    service: ProductService = Depends(get_product_service),
):
    updated_product = await service.update_product(product_id, product_data)
    logger.info(f"Product updated: {updated_product.name}")
    return updated_product


@router.delete("/{product_id}")
async def delete_product(
    product_id: int, service: ProductService = Depends(get_product_service)
):
    deleted_product = await service.delete_product(product_id)
    logger.info(f"Product deleted: {deleted_product.name}")
    return {"detail": f"Product {deleted_product.name} deleted"}
